"""Excel export service."""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import List
from utils.helpers import format_currency, format_quarter


class ExportService:
    """Service for exporting data to Excel."""
    
    @staticmethod
    def export_allocation_report(
        expense_data: dict,
        allocations: List[dict],
        output_path: str
    ) -> bool:
        """
        Export allocation report to Excel with formatting.
        
        Args:
            expense_data: Dictionary with expense information
            allocations: List of allocation dictionaries
            output_path: Path to save Excel file
        
        Returns:
            bool: Success status
        """
        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Expense Details
                expense_df = pd.DataFrame([{
                    'Số tài khoản': expense_data.get('account_number'),
                    'Tên khoản mục': expense_data.get('name'),
                    'Mã chứng từ': expense_data.get('document_code', ''),
                    'Tổng tiền': expense_data.get('total_amount'),
                    'Ngày bắt đầu': expense_data.get('start_date'),
                    'Ngày kết thúc': expense_data.get('end_date'),
                    'Mã phụ': expense_data.get('sub_code'),
                    'Số tháng phân bổ': expense_data.get('allocation_months')
                }])
                expense_df.to_excel(writer, sheet_name='Thông tin chi phí', index=False)
                
                # Sheet 2: Allocation Schedule
                allocation_records = []
                cumulative_allocated = 0
                total_amount = expense_data.get('total_amount')
                
                for alloc in allocations:
                    cumulative_allocated += alloc['amount']
                    remaining_balance = total_amount - cumulative_allocated
                    
                    allocation_records.append({
                        'Quý': format_quarter(alloc['quarter'], alloc['year']),
                        'Năm': alloc['year'],
                        'Ngày bắt đầu': alloc['start_date'].strftime('%d/%m/%Y'),
                        'Ngày kết thúc': alloc['end_date'].strftime('%d/%m/%Y'),
                        'Số ngày': alloc['days_in_quarter'],
                        'Số tiền phân bổ': alloc['amount'],
                        'Tỷ lệ (%)': round((alloc['days_in_quarter'] / alloc['total_days']) * 100, 2),
                        'Lũy kế phân bổ': cumulative_allocated,
                        'Còn lại': remaining_balance
                    })
                
                allocation_df = pd.DataFrame(allocation_records)
                allocation_df.to_excel(writer, sheet_name='Kế hoạch phân bổ', index=False)
                
                # Add summary row
                summary_row = pd.DataFrame([{
                    'Quý': 'TỔNG CỘNG',
                    'Năm': '',
                    'Ngày bắt đầu': '',
                    'Ngày kết thúc': '',
                    'Số ngày': sum(a['days_in_quarter'] for a in allocations),
                    'Số tiền phân bổ': sum(a['amount'] for a in allocations),
                    'Tỷ lệ (%)': 100.00,
                    'Lũy kế phân bổ': total_amount,
                    'Còn lại': 0
                }])
                summary_row.to_excel(
                    writer,
                    sheet_name='Kế hoạch phân bổ',
                    startrow=len(allocation_df) + 1,
                    index=False,
                    header=False
                )
            
            # Apply formatting
            ExportService._format_excel(output_path)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            return False
    
    @staticmethod
    def _format_excel(file_path: str):
        """Apply formatting to Excel file."""
        try:
            wb = load_workbook(file_path)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Border for all cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Format summary row if it's the allocation sheet
                if sheet_name == 'Kế hoạch phân bổ':
                    summary_row = ws.max_row
                    summary_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    summary_font = Font(bold=True)
                    
                    for cell in ws[summary_row]:
                        cell.fill = summary_fill
                        cell.font = summary_font
            
            wb.save(file_path)
            
        except Exception as e:
            print(f"Error formatting Excel: {str(e)}")
    
    @staticmethod
    def export_multiple_expenses(
        expenses_data: List[dict],
        output_path: str
    ) -> bool:
        """
        Export multiple expenses to a single Excel file.
        
        Args:
            expenses_data: List of expense dictionaries with allocations
            output_path: Path to save Excel file
        
        Returns:
            bool: Success status
        """
        try:
            all_records = []
            
            for expense in expenses_data:
                for alloc in expense.get('allocations', []):
                    all_records.append({
                        'Số tài khoản': expense.get('account_number'),
                        'Tên khoản mục': expense.get('name'),
                        'Mã phụ': expense.get('sub_code'),
                        'Quý': format_quarter(alloc['quarter'], alloc['year']),
                        'Năm': alloc['year'],
                        'Ngày bắt đầu': alloc['start_date'].strftime('%d/%m/%Y'),
                        'Ngày kết thúc': alloc['end_date'].strftime('%d/%m/%Y'),
                        'Số ngày': alloc['days_in_quarter'],
                        'Số tiền phân bổ': alloc['amount']
                    })
            
            df = pd.DataFrame(all_records)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Tổng hợp phân bổ', index=False)
            
            ExportService._format_excel(output_path)
            
            return True
            
        except Exception as e:
            print(f"Error exporting multiple expenses: {str(e)}")
            return False
